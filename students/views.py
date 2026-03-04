from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Student, Class, Dorm, Form
from .forms import StudentForm, ClassForm, DormForm, FormLevelForm


def get_school(request):
    """Helper — returns the school of the logged-in user"""
    return request.user.school


# ── STUDENT LIST ──────────────────────────────────────────
@login_required
def student_list(request):
    school   = get_school(request)
    students = Student.objects.filter(school=school).select_related('current_class', 'dorm')
    return render(request, 'students/list.html', {'students': students})


# ── STUDENT CREATE ─────────────────────────────────────────
@login_required
def student_create(request):
    school = get_school(request)
    form   = StudentForm(school, request.POST or None)
    if form.is_valid():
        student        = form.save(commit=False)
        student.school = school
        student.save()
        messages.success(request, f"Student {student.get_full_name()} added successfully.")
        return redirect('student_list')
    return render(request, 'students/form.html', {'form': form, 'title': 'Add Student'})


# ── STUDENT UPDATE ─────────────────────────────────────────
@login_required
def student_update(request, pk):
    school  = get_school(request)
    student = get_object_or_404(Student, pk=pk, school=school)
    form    = StudentForm(school, request.POST or None, instance=student)
    if form.is_valid():
        form.save()
        messages.success(request, "Student updated successfully.")
        return redirect('student_list')
    return render(request, 'students/form.html', {'form': form, 'title': 'Edit Student'})


# ── STUDENT DELETE ─────────────────────────────────────────
@login_required
def student_delete(request, pk):
    school  = get_school(request)
    student = get_object_or_404(Student, pk=pk, school=school)
    if request.method == 'POST':
        student.status = Student.STATUS_INACTIVE
        student.save()
        messages.success(request, "Student deactivated successfully.")
        return redirect('student_list')
    return render(request, 'students/confirm_delete.html', {'student': student})


# ── STUDENT DETAIL ─────────────────────────────────────────
@login_required
def student_detail(request, pk):
    school  = get_school(request)
    student = get_object_or_404(Student, pk=pk, school=school)
    history = student.class_history.all().order_by('-changed_at')
    return render(request, 'students/detail.html', {'student': student, 'history': history})

from .imports import import_students_from_xlsx, import_students_from_csv


# ── IMPORT STUDENTS ───────────────────────────────────────
@login_required
def student_import(request):
    if not request.user.is_school_admin():
        messages.error(request, "Only admins can import students.")
        return redirect('student_list')

    result = None
    if request.method == 'POST' and request.FILES.get('import_file'):
        file      = request.FILES['import_file']
        school    = get_school(request)
        filename  = file.name.lower()

        if filename.endswith('.xlsx'):
            result = import_students_from_xlsx(file, school)
        elif filename.endswith('.csv'):
            result = import_students_from_csv(file, school)
        else:
            messages.error(request, "Only .xlsx and .csv files are supported.")
            return redirect('student_import')

        if result['success'] or result['updated']:
            messages.success(
                request,
                f"{result['success']} students imported, {result['updated']} updated."
            )

    return render(request, 'students/import.html', {'result': result})


# ── DOWNLOAD IMPORT TEMPLATE ──────────────────────────────
@login_required
def download_import_template(request):
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Headers
    headers = [
        'admission_number', 'first_name', 'last_name',
        'class', 'dorm', 'date_admitted', 'status'
    ]
    ws.append(headers)

    # Example rows
    ws.append(['ADM001', 'John', 'Kamau', 'Form 1K', 'Kibaki Dorm', '2024-01-15', 'active'])
    ws.append(['ADM002', 'Mary', 'Wanjiru', 'Form 2E', 'Uhuru Dorm', '2024-01-15', 'active'])

    # Style header row
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1a1a2e')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    wb.save(response)
    return response


# ── BULK DELETE ───────────────────────────────────────────
@login_required
def student_bulk_delete(request):
    if not request.user.is_school_admin():
        messages.error(request, "Only admins can perform bulk delete.")
        return redirect('student_list')

    school = get_school(request)

    if request.method == 'POST':
        confirm_text = request.POST.get('confirm_text', '').strip()
        reason       = request.POST.get('reason', '').strip()
        student_ids  = request.POST.getlist('student_ids')

        # Security check
        if confirm_text != 'DELETE':
            messages.error(request, "You must type DELETE to confirm.")
            return redirect('student_bulk_delete')

        if not reason:
            messages.error(request, "You must provide a reason for bulk delete.")
            return redirect('student_bulk_delete')

        if not student_ids:
            messages.error(request, "No students selected.")
            return redirect('student_bulk_delete')

        # Soft delete — set inactive
        count = Student.objects.filter(
            school=school,
            pk__in=student_ids
        ).update(status=Student.STATUS_INACTIVE)

        # Log the action
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(
            f"BULK DELETE: {request.user} deactivated {count} students. Reason: {reason}"
        )

        messages.success(request, f"{count} students deactivated successfully.")
        return redirect('student_list')

    # GET — show bulk delete page
    students = Student.objects.filter(
        school=school
    ).exclude(
        status=Student.STATUS_INACTIVE
    ).select_related('current_class', 'dorm')

    return render(request, 'students/bulk_delete.html', {'students': students})