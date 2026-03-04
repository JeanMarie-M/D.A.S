from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from .models import DutyArea, DutyAssignment, DutySwapRequest
from .forms import DutyAreaForm, ManualAssignForm, SwapRequestForm
from .engine import allocate_duties
from schools.models import Term
from .imports import import_duties_from_xlsx, import_duties_from_csv

@login_required
def duty_area_import(request):
    if not request.user.is_school_admin():
        messages.error(request, "Access denied.")
        return redirect('duty_area_list')

    school = request.user.school
    result = None

    if request.method == 'POST' and request.FILES.get('import_file'):
        file     = request.FILES['import_file']
        filename = file.name.lower()

        if filename.endswith('.xlsx'):
            result = import_duties_from_xlsx(file, school)
        elif filename.endswith('.csv'):
            result = import_duties_from_csv(file, school)
        else:
            messages.error(request, "Only .xlsx and .csv files are supported.")
            return redirect('duty_area_import')

        if result['success'] or result['updated']:
            messages.success(
                request,
                f"✅ {result['success']} duty areas imported, "
                f"{result['updated']} updated."
            )

    return render(request, 'duties/import.html', {'result': result})


@login_required
def duty_import_template(request):
    """Download a blank Excel template for duty area import."""
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duty Areas"

    from openpyxl.styles import Font, PatternFill, Alignment
    headers = [
        'label', 'name', 'area_size', 'is_heavy',
        'students_required', 'specialization',
        'specific_form', 'specific_dorm', 'tools_required'
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF', name='Arial')
        cell.fill      = PatternFill('solid', start_color='1a1a2e')
        cell.alignment = Alignment(horizontal='center')

    # Example rows
    examples = [
        ['NHS-DH', 'Dining Hall Cleaning', 'large',  'yes', 8, 'none', '', '', 'Mops, Brooms'],
        ['NHS-KT', 'Kitchen Duty',         'large',  'yes', 6, 'none', '', '', 'Aprons, Gloves'],
        ['NHS-FA', 'Form 1 Classrooms',    'medium', 'no',  4, 'form', 'Form 1', '', 'Brooms, Dusters'],
        ['NHS-DA', 'Uhuru Dorm Cleaning',  'medium', 'no',  5, 'dorm', '', 'Uhuru Dorm', 'Mops, Buckets'],
        ['NHS-LB', 'Library Duty',         'small',  'no',  3, 'none', '', '', 'Dusters'],
    ]
    for row in examples:
        ws.append(row)

    # Column widths
    widths = [12, 25, 10, 10, 18, 14, 14, 14, 20]
    for i, w in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="duty_areas_template.xlsx"'
    wb.save(response)
    return response

def get_school(request):
    return request.user.school


# ── DUTY AREAS ────────────────────────────────────────────
@login_required
def duty_area_list(request):
    school = get_school(request)
    areas  = DutyArea.objects.filter(school=school, is_active=True)
    return render(request, 'duties/area_list.html', {'areas': areas})


@login_required
def duty_area_create(request):
    school = get_school(request)
    form   = DutyAreaForm(school, request.POST or None)
    if form.is_valid():
        area        = form.save(commit=False)
        area.school = school
        area.save()
        messages.success(request, f"Duty area '{area}' created.")
        return redirect('duty_area_list')
    return render(request, 'duties/area_form.html', {'form': form, 'title': 'Add Duty Area'})


@login_required
def duty_area_update(request, pk):
    school = get_school(request)
    area   = get_object_or_404(DutyArea, pk=pk, school=school)
    form   = DutyAreaForm(school, request.POST or None, instance=area)
    if form.is_valid():
        form.save()
        messages.success(request, "Duty area updated.")
        return redirect('duty_area_list')
    return render(request, 'duties/area_form.html', {'form': form, 'title': 'Edit Duty Area'})


# ── AUTO ALLOCATE ─────────────────────────────────────────
@login_required
def allocate_view(request):
    if request.user.role not in ['admin', 'prefect', 'superadmin']:
        messages.error(request, "Access denied.")
        return redirect('dashboard')

    school       = get_school(request)
    current_term = Term.objects.filter(school=school, is_current=True).first()

    if not current_term:
        messages.error(request, "No active term set. Please create a term first.")
        return redirect('term_create')

    duty_areas = DutyArea.objects.filter(school=school, is_active=True)

    if not duty_areas.exists():
        messages.error(request, "No duty areas found. Please add duty areas first.")
        return redirect('duty_area_list')

    if request.method == 'POST':
        rotation = int(request.POST.get('rotation', 1))
        try:
            result = allocate_duties(school, current_term, rotation, assigned_by=request.user)
            assigned = result.get('assigned', [])
            count = len(assigned) if isinstance(assigned, list) else assigned
            messages.success(
                request,
                f" Allocation complete! {count} students assigned."
            )
        except Exception as e:
            messages.error(request, f"Allocation failed: {e}")
        return redirect('allocation_summary')
    # GET — calculate next rotation number
    from students.models import Student
    eligible     = Student.objects.filter(school=school, status='active').count()
    last         = DutyAssignment.objects.filter(
                       school=school, term=current_term
                   ).order_by('-rotation').first()
    next_rotation = (last.rotation + 1) if last else 1

    return render(request, 'duties/allocate.html', {
        'term':          current_term,
        'duty_areas':    duty_areas,
        'eligible':      eligible,
        'next_rotation': next_rotation,
    })


# ── ALLOCATION SUMMARY ────────────────────────────────────
@login_required
def allocation_summary(request):
    school       = get_school(request)
    current_term = Term.objects.filter(school=school, is_current=True).first()

    if not current_term:
        messages.error(request, "No active term found.")
        return redirect('dashboard')

    # Get selected rotation or default to latest
    rotation = request.GET.get('rotation')
    if rotation:
        rotation = int(rotation)
    else:
        last = DutyAssignment.objects.filter(
            school=school, term=current_term
        ).order_by('-rotation').first()
        rotation = last.rotation if last else 1

    assignments = DutyAssignment.objects.filter(
        school=school,
        term=current_term,
        rotation=rotation,
    ).select_related('student', 'duty_area').order_by('duty_area__label')

    # Get all rotations for this term for the dropdown
    rotations = DutyAssignment.objects.filter(
        school=school, term=current_term
    ).values_list('rotation', flat=True).distinct().order_by('rotation')

    # Group by duty area
    from collections import defaultdict
    grouped = defaultdict(list)
    for a in assignments:
        grouped[a.duty_area].append(a.student)

    return render(request, 'duties/allocation_summary.html', {
        'assignments':  assignments,
        'grouped':      dict(grouped),
        'current_term': current_term,
        'rotation':     rotation,
        'rotations':    rotations,
        'total':        assignments.count(),
    })

# ── MANUAL ASSIGN ─────────────────────────────────────────
@login_required
def manual_assign(request):
    if not request.user.is_school_admin():
        messages.error(request, "Only admins can manually assign duties.")
        return redirect('duty_area_list')

    school = get_school(request)
    term   = Term.objects.filter(school=school, is_current=True).first()
    form   = ManualAssignForm(school, term, request.POST or None)

    if form.is_valid():
        assignment             = form.save(commit=False)
        assignment.school      = school
        assignment.term        = term
        assignment.method      = 'manual'
        assignment.assigned_by = request.user
        assignment.save()
        messages.success(request, "Duty manually assigned.")
        return redirect('duty_area_list')

    return render(request, 'duties/manual_assign.html', {'form': form})


# ── SWAP REQUEST ──────────────────────────────────────────
@login_required
def swap_request(request):
    school = get_school(request)
    term   = Term.objects.filter(school=school, is_current=True).first()
    form   = SwapRequestForm(school, term, request.POST or None)

    if form.is_valid():
        swap              = form.save(commit=False)
        swap.school       = school
        swap.term         = term
        swap.rotation     = int(request.POST.get('rotation', 1))
        swap.requested_by = request.user
        swap.save()
        messages.success(request, "Swap request submitted. Awaiting admin approval.")
        return redirect('swap_list')

    return render(request, 'duties/swap_form.html', {'form': form})


@login_required
def swap_list(request):
    school = get_school(request)
    swaps  = DutySwapRequest.objects.filter(school=school).order_by('-id')
    return render(request, 'duties/swap_list.html', {'swaps': swaps})


@login_required
def swap_review(request, pk):
    if not request.user.is_school_admin():
        messages.error(request, "Only admins can approve swaps.")
        return redirect('swap_list')

    swap = get_object_or_404(DutySwapRequest, pk=pk)

    if request.method == 'POST':
        action           = request.POST.get('action')
        swap.reviewed_by = request.user
        swap.reviewed_at = timezone.now()
        swap.admin_note  = request.POST.get('admin_note', '')

        if action == 'approve':
            try:
                a1 = DutyAssignment.objects.get(
                    term=swap.term, student=swap.from_student,
                    rotation=swap.rotation
                )
                a2 = DutyAssignment.objects.get(
                    term=swap.term, student=swap.to_student,
                    rotation=swap.rotation
                )
                a1.duty_area, a2.duty_area = a2.duty_area, a1.duty_area
                a1.method = a2.method = 'swap'
                a1.save()
                a2.save()
                swap.status = 'approved'
                messages.success(request, "Swap approved and executed.")
            except DutyAssignment.DoesNotExist:
                messages.error(request, "Could not find assignments to swap.")
                swap.status = 'rejected'
        else:
            swap.status = 'rejected'
            messages.warning(request, "Swap rejected.")

        swap.save()
        return redirect('swap_list')

    return render(request, 'duties/swap_review.html', {'swap': swap})


# ── DUTY AREA IMPORT ──────────────────────────────────────
@login_required
def duty_area_import(request):
    if not request.user.is_school_admin():
        messages.error(request, "Access denied.")
        return redirect('duty_area_list')

    school = get_school(request)
    result = None

    if request.method == 'POST' and request.FILES.get('import_file'):
        from .imports import import_duties_from_xlsx, import_duties_from_csv
        file     = request.FILES['import_file']
        filename = file.name.lower()

        if filename.endswith('.xlsx'):
            result = import_duties_from_xlsx(file, school)
        elif filename.endswith('.csv'):
            result = import_duties_from_csv(file, school)
        else:
            messages.error(request, "Only .xlsx and .csv files are supported.")
            return redirect('duty_area_import')

        if result['success'] or result['updated']:
            messages.success(
                request,
                f"✅ {result['success']} duty areas imported, "
                f"{result['updated']} updated."
            )

    return render(request, 'duties/import.html', {'result': result})


@login_required
def duty_import_template(request):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duty Areas"

    headers = [
        'label', 'name', 'area_size', 'is_heavy',
        'students_required', 'specialization',
        'specific_form', 'specific_dorm', 'tools_required'
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF', name='Arial')
        cell.fill      = PatternFill('solid', start_color='1a1a2e')
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ['NHS-DH', 'Dining Hall Cleaning', 'large',  'yes', 8, 'none', '', '', 'Mops, Brooms'],
        ['NHS-KT', 'Kitchen Duty',         'large',  'yes', 6, 'none', '', '', 'Aprons, Gloves'],
        ['NHS-FA', 'Form 1 Classrooms',    'medium', 'no',  4, 'form', 'Form 1', '', 'Brooms, Dusters'],
        ['NHS-DA', 'Uhuru Dorm Cleaning',  'medium', 'no',  5, 'dorm', '', 'Uhuru Dorm', 'Mops, Buckets'],
        ['NHS-LB', 'Library Duty',         'small',  'no',  3, 'none', '', '', 'Dusters'],
    ]
    for row in examples:
        ws.append(row)

    widths = [12, 25, 10, 10, 18, 14, 14, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="duty_areas_template.xlsx"'
    wb.save(response)
    return response